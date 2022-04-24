import xlrd
import json
import openpyxl
from common.file.handle_system import adjust_path_data


def is_josn(inStr):
    try:
        json.loads(inStr)
    except:#只有try报错才执行里面的代码
        return False
    return True

def is_dict(inStr):
    try:
        type(inStr) == dict
    except:#只有try报错才执行里面的代码
        return False
    return True



def get_excel_data(excelPath,sheetIndex):
    """
    读取excel文件的方法
    :param excelPath: 文件路径
    :param sheetName: sheet页的名称
    :param caseName: 执行的case名称
    :return:
    """
    # 存放字典格式，{sheet1页的名称:[[第一行数据],[第二行数据][第三行数据]],sheet2页的名称:[[第一行数据],[第二行数据][第三行数据]]}
    sheet_dict = {}
    excelPath=adjust_path_data(excelPath)
    # 打开excel文件，并且保持原有样式
    workBook = xlrd.open_workbook(filename=excelPath,formatting_info=True)
    # Frame().logger(fileLog=True).info(f'获取用例路径为:{excelPath}的文件')
    # 拿到sheet页的内容
    table = workBook.sheet_by_index(sheetIndex)
    # 存放列表格式,，[[第一行数据],[第二行数据][第三行数据]],sheet2页的名称:[[第一行数据],[第二行数据][第三行数据]]
    sheet_datas = []
    # print(table)
    # table.nrows是总行数，按行号进行遍历，(从第一行开始，总行数)
    for row in range(1,table.nrows):
        # 定义一个空列表，存放的是每一行的数据
        row_data = []
        # 得到case状态
        cell_data = table.cell_value(row,4)
        # 如果case状态为无效，就跳过此次循环，不将此行数据加入row_data
        if cell_data =='否':
            continue
        # table.ncols是总列数，按列号进行遍历，(从第0列开始，总列)

        for col in range(0,table.ncols):
            # 获取指定单元格数据(行，列)
            cell_data = table.cell_value(row,col)
            # 如果是请求头、请求参数和响应参数这三列数据
            if col in (6,7):
                try:
                    cell_data = cell_data.replace('\n','').replace(' ','')
                    if is_josn(cell_data):  # 判断当前单元格数据是否是json
                        cell_data = json.loads(cell_data)  # 转化为字典格式
                        # print(f'cell_data:{cell_data}的类型是：{type(cell_data)}')
                except Exception as error:
                    raise error
            if col in (3,5):
                cell_data = cell_data.split('\n')

            # 将当前行所有单元格的数据追加进行数据列表

            row_data.append(cell_data)
            if col in (4,):
                row_data.remove(cell_data)
        # 将当前行列表数据追加到sheet页数据列表
        sheet_datas.append(row_data)
    return sheet_datas
    # Frame().logger(fileLog=True).info(f'获取指定sheet页：{sheetName}的所有行数据')


def writExcle(filePath,exlSheetName,dataLists,exlRow,exlColumn):
    """
    filePath:文件路径+文件名（后缀支持.xlsx），举例："C:/Users/qinyan5/Desktop/bagtestData.xlsx"
    exlSheetName：插入表格sheet页name
    dataLists：输入的数据data，格式：[[],[],[]]，举例：[['aaa','bbb','ccc'],['ggg','hhh','kkk'],['11','22','333']]
    需要的写入第几行，第几列
    exlRow:从表格的第几行开始输入，坐标：行
    exlColumn：从表格的第几列开始输入，坐标：列
    """
    filePath = adjust_path_data(filePath)
    dataListlen = len(dataLists)
    datalen = len(dataLists[0])
    wb = openpyxl.load_workbook(filePath)
    sh = wb[exlSheetName]
    for i in range(0,dataListlen):
        data = dataLists[i]
        for j in range(0,datalen):
            sh.cell(i+exlRow, j+exlColumn, data[j])
    wb.save(filePath)

#根据excle单元格列坐标批量读出数据
def readExcle(filePath,sheetName,columStart,columnX):
    """
    :param filePath: 文件路径
    :param sheetName: excle sheet页名称
    :param columStart: 开始取列码 :B2
    :param columnX: 列标记:B
    :return:
    """

    #打开excle文件
    filePath = adjust_path_data(filePath)
    wb = openpyxl.load_workbook(filePath)
    #sheet页
    sh = wb[sheetName]
    #获取最大列表数据行数
    max_row = (sh.max_row) -1
    # print(max_row)
    #获取需要去字段的单元格区间:sh[B2:B7]
    columnX = sh[columStart:columnX+'%d' % max_row]

    columnValueList = []
    #获取该列对应的所有的单元格对象
    for column_cells in columnX:
        for cell in column_cells:
            columnvalue = cell.value
            columnValueList.append(columnvalue)
    # print(columnValueList)
    return columnValueList


def readAllList(filePath,sheetName):
    """
    取全表数据，第一行作为title，每一行作为value
    :param filePath:
    :param sheetName:
    :return:
    """
    filePath = adjust_path_data(filePath)
    wb = openpyxl.load_workbook(filePath)
    sh = wb[sheetName]
    #表格第一行内容为title
    title = [i.value for i in sh[1]]
    max_row = sh.max_row + 1

    #key-->value
    data_list = []
    for j in range(2, max_row):
        data = [''if i.value == None else i.value for i in sh[j] ]
        datadict = dict(zip(title,data))
        data_list.append(datadict)
    return data_list

def readValueList(filePath,sheetName):
    """
    取全部sheet页数据，除第一行，其他作为list value
    :param filePath:
    :param sheetName:
    :return:
    """
    filePath = adjust_path_data(filePath)
    wb = openpyxl.load_workbook(filePath)
    sh = wb[sheetName]

    max_row = sh.max_row

    dataAll_list = []
    for i in range(2,max_row):
        data = ['' if j.value == None else j.value for j in sh[i]]

        dataAll_list.append(data)
    return dataAll_list




def readExcleMerge(filePath,sheetName):
    """
    取全表合并单元格坐标
    :param filePath:
    :param sheetName:
    :return:
    """
    filePath = adjust_path_data(filePath)
    wb = openpyxl.load_workbook(filePath)
    sh = wb[sheetName]
    exlSpace = sh.merged_cells.ranges

    for spaceData in exlSpace:
        datavalue = spaceData.start_cell.value
        # print(datavalue)
        return datavalue

def excel_to_list(data_file, sheet,_index:int=1):
    """
    读取Excel中特定sheet的数据，按行将数据存入数组datalist
    :param data_file:Excel文件目录
    :param sheet:需要读取的sheet名称
    :return:datalist
    """
    data_file = adjust_path_data(data_file)
    # 新建个空列表，来存储所有的数据
    data_list = []
    # 打开excel
    wb = xlrd.open_workbook(data_file)
    # 获取工作簿
    sh = wb.sheet_by_name(sheet)
    # 获取标题行数据
    header = sh.row_values(0)
    # 跳过标题行，从第二行开始取数据
    for i in range(_index, sh.nrows):
        # 将标题和每行数据组装成字典
        d = dict(zip(header, sh.row_values(i)))
        # 列表嵌套字典格式，每个元素是一个字典
        data_list.append(d)
    return data_list


def get_test_data(data_list, case_name):
    """
    根据用例名称查找用例本身
    :param data_list:按行存储用例的数组
    :param case_name: 用例名称
    :return: case_data
    """
    for case_data in data_list:
        # 如果字典数据中case_name与参数一致
        if case_name == case_data['用例标题']:
            return case_data
        # 如果查询不到会返回None
        else:
            return None




if __name__ == '__main__':
    # readList = readAllList(filePath, "arrivalBindCase")
    # print(readList)
    print(excel_to_list('/Users/edz/Desktop/aa.xls','Sheet1',12))
    # sheet_data= excel_to_list(CONFIG_PATH+'\\case_data.xls','Sheet1')
    # sheet_data2 = get_test_data(sheet_data, '案件编号(caseId)为空,上传提交失败')
    # print(sheet_data2)
    # for data in sheet_data:
    #     print(data['用例编号'])
